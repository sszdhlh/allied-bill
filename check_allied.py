import openpyxl
from openpyxl.styles import PatternFill, Font
import csv
import re
import sys

# HD/HOME DELIVERY
# DHS/DEPOT HANDLING SURCHARGE
# WS/WIDTH SURCHARGE
# MHF/MANUAL HANDLING FEE
# MM/MANUAL MEASURING FEE
# Includes Out Of Area charges
# 2MC/ENT MIN 2 MAN CREW
# RTS/RETURN TO SENDER
# RD/RE DELIVERIES
# LSC/LENGTH SURCHARGE 
# CB/CALL BEFORE
fee_code = ["HD/", "DHS/", "WS/", "MHF/", "MM/", "2MC/", "RTS/", "RD/", "LSC/", "CB/", "TL/"]
fee_names = ["HOME DELIVERY","DEPOT HANDLING SURCHARGE","WIDTH SURCHARGE","MANUAL HANDLING FEE","MANUAL MEASURING FEE","ENT MIN 2 MAN CREW", "RETURN TO SENDER", "RD DELIVERIES", "LENGTH SURCHARGE", "CALL BEFORE","TAIL LIFT REQUIRED"]
out_of_area = "INCLUDES OUT OF AREA CHARGES"

workbook = openpyxl.load_workbook("template.xlsx")
sheet = workbook["Total Bill"]

colory = ["ffeb9c", "9c6500"]
filly = PatternFill('solid', fgColor=colory[0])

colorg = ["c6efce", "006100"]
fillg = PatternFill('solid', fgColor=colorg[0])
fontg = Font('Arial', color = colory[1], size=12)

order_list = []
route_list = []
duplicated_route = []
location_list = []
charge_list = []
count_list = []
total_surcharge_list = []
fee_dict_list = []
fuel_surcharge_list = []

row_num = sheet.max_row
for i in range(row_num):
    current_detail = sheet.cell(row=i+2, column =4).value
    if current_detail != None and current_detail != "" and current_detail.strip().upper() != "TRANSACTION DETAILS":
        if sheet.cell(row=i+2, column =1).value == None or  sheet.cell(row=i+2, column =1).value == "":
            sheet.cell(row=i+1, column =4).value += "\n" + current_detail
            sheet.cell(row=i+2, column =4).value = ""

workbook.save("template.xlsx")

workbook = openpyxl.load_workbook("template.xlsx")
sheet = workbook["Total Bill"]
for row in sheet.iter_rows(min_row=2):
    # print(row[5].value)
    if row[3].value is None and row[0].value is not None:
        if str(row[0].value).strip() == "Fuel Surcharge":
        # 执行代码

            fuel_surcharge_list.append(row[5].value)
            continue
    if row[3].value != None and row[3].value != "" and row[3].value.strip().upper() != "TRANSACTION DETAILS":
        
        details = row[3].value.split("\n")
        # print(details[0])
        if details[0].strip() not in route_list:
            route_list.append(details[0])
        else:
            duplicated_route.append(details[0])
        locations = details[0].split(" - ")
        # if locations[0].strip() not in location_list:
        #     location_list.append(locations[0].strip())
        #     charge_list.append(float(row[5].value))
        # else:
        #     index = location_list.index(locations[0].strip())
        #     charge_list[index] += float(row[5].value)

        if locations[1].strip() not in location_list:
            location_list.append(locations[1].strip())
            charge_list.append(float(row[5].value))
            count_list.append(1)
        else:
            index = location_list.index(locations[1].strip())
            charge_list[index] += float(row[5].value)
            count_list[index] += 1
        
        detail_str = row[3].value.replace(details[0].strip(),"")
        surcharge_list = re.findall(r"\$\d+\.?\d*", detail_str)
        if len(surcharge_list) > 0:
            for i in range(len(surcharge_list)):
                surcharge_list[i] = float(surcharge_list[i].replace("$", ""))
            
            total_surcharge = sum(surcharge_list)
        else:
            total_surcharge = 0

        total_surcharge_list.append(total_surcharge)

        row[6].value = total_surcharge

        fee_dict = {
            "HD/": 0,
            "DHS/": 0,
            "WS/": 0,
            "MHF/": 0, 
            "MM/": 0,
            "2MC/": 0,
            "RTS/": 0,
            "RD/": 0, 
            "LSC/": 0, 
            "CB/": 0,
            "TL/": 0,
        }

        column_fee_dict = {
            "HD/": 9,
            "DHS/": 10,
            "WS/": 11,
            "MHF/": 12, 
            "MM/": 13,
            "2MC/": 14,
            "RTS/": 15,
            "RD/": 16, 
            "LSC/": 17, 
            "CB/": 18,
            "TL/":19
        }
        new_detail_str = ""
        for fee in fee_code:
            fee_count = 0
            while (fee in detail_str):
                fee_count += 1
                start_index = detail_str.index(fee)
                new_detail_str = detail_str[start_index:]
                fee_start_index = new_detail_str.index("($")
                fee_end_index = new_detail_str.index(")")
                fees = new_detail_str[fee_start_index+2: fee_end_index]
                # print(float(fees))
                detail_str = detail_str.replace(fee, "", 1)
                detail_str = detail_str.replace("($"+fees+")", "", 1)

                fee_dict[fee] += float(fees)
            
            # print(new_detail_str)
            if fee_count >=2:
                row[column_fee_dict[fee]].fill = filly
            
            row[column_fee_dict[fee]].value = fee_dict[fee]
        fee_dict_list.append(fee_dict)

        if out_of_area.upper() in row[3].value.upper():
            row[8].value = "True"

        if "DIMENSION" in row[3].value.upper():
            row[7].value = "True"
            row[7].fill = fillg
        
        # sys.exit()

for row in sheet.iter_rows(min_row=2):
    if row[3].value != None and row[3].value != "":
        details = row[3].value.split("\n")
        if details[0].strip() in duplicated_route:
            row[3].font = fontg
    
        
        

summary_sheet = workbook["Summary"]
for i in range(len(location_list)):
    # print(location_list[i])
    suburb = location_list[i].split(" ")[0]
    if re.findall('\d{4}', location_list[i]) != []:
        postcode = re.findall('\d{4}', location_list[i])[0]
    else:
        suburb = location_list[i]
        postcode = ""
    summary_sheet.cell(row=i+2, column=1).value = suburb
    summary_sheet.cell(row=i+2, column=2).value = postcode
    summary_sheet.cell(row=i+2, column=3).value = charge_list[i]
    summary_sheet.cell(row=i+2, column=4).value = count_list[i]

fee_total_dict = {
    "HD/": 0,
    "DHS/": 0,
    "WS/": 0,
    "MHF/": 0, 
    "MM/": 0,
    "2MC/": 0,
    "RTS/": 0,
    "RD/": 0, 
    "LSC/": 0, 
    "CB/": 0,
    "TL/": 0,
}

for fee_dict in fee_dict_list:
    for key, value in fee_dict.items():
        if key in fee_total_dict:
            fee_total_dict[key] += float(value)

i = 2
for key, value in fee_total_dict.items():
    index = fee_code.index(key)
    fee_name = fee_names[index]
    summary_sheet.cell(row=i, column=6).value = fee_name
    summary_sheet.cell(row=i, column=7).value = value
    i += 1

summary_sheet.cell(row=i, column =6).value = "Fuel Surcharge"
summary_sheet.cell(row=i, column =7).value = sum(fuel_surcharge_list)

workbook.save("week35.xlsx")


print(fuel_surcharge_list)
